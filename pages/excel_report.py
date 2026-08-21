from io import BytesIO

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font

from pages.utils import parse_total_marks


def load_data(path):
    return st.session_state.stored_data.get(path, [])


def create_excel_sheet():
    data2 = load_data("Result_dict")
    if not data2:
        st.warning("No detailed data available. Please process a PDF first.")
        return

    with st.spinner("Creating Excel sheet..."):
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Results"
        all_codes = []
        seen = set()
        for student in data2:
            for code in student.get("Code", []):
                if isinstance(code, str) and code not in seen:
                    seen.add(code)
                    all_codes.append(code)

        header = ["Seat No", "Name"]
        for code in all_codes:
            header.extend([code, "UA", "CA", "Total", "Subject_Status"])
        header.extend(["Total Marks", "Status", "Percentage"])
        ws.append(header)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for student in data2:
            row = [student.get("Seat No", ""), student.get("Name", "")]
            codes = student.get("Code", [])
            ua_list = student.get("UA", [])
            ca_list = student.get("CA", [])
            total_list = student.get("Total", [])
            status_list = student.get("Status1", [])
            lookup = {}
            for i, code in enumerate(codes):
                if not isinstance(code, str):
                    continue
                lookup[code] = (
                    ua_list[i] if i < len(ua_list) else "",
                    ca_list[i] if i < len(ca_list) else "",
                    total_list[i] if i < len(total_list) else "",
                    status_list[i] if i < len(status_list) else "",
                )

            total_val = sum(parse_total_marks(v) for v in total_list[:9])
            status = "Fail" if "F" in status_list[:16] else "Pass"
            for code in all_codes:
                if code in lookup:
                    ua, ca, total, subj_status = lookup[code]
                    row.extend(["", ua, ca, total, subj_status])
                else:
                    row.extend(["", "", "", "", ""])

            percentage = f"{(total_val / 900) * 100:.2f}" if total_list else "0.00"
            row.extend([total_val, status, percentage])
            ws.append(row)
            current_row = ws.max_row
            ws[f"A{current_row}"].font = Font(bold=True)

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        st.success("Excel sheet created successfully!")
        st.download_button(
            label="📥 Download Excel File",
            data=excel_buffer,
            file_name="BCS-II_Results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def show():
    st.header("📝 Generate Detailed Excel Report")
    st.info("This will create a comprehensive Excel sheet with all student marks.")
    if st.button("Generate Excel"):
        create_excel_sheet()
